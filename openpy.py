from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Alignment

"""
CREAR UN NUEVO ARCHIVO OPENPYXL

"""
from openpyxl import Workbook
import time
book = Workbook()
sheet = book.active
sheet['A1'] = 56
sheet['A2'] = 43
now = time.strftime("%x")
sheet['A3'] = now
book.save("sample.xlsx")

"""
ESCRIBIR EN UNA CELDA CON OPENPYXL
"""
book = Workbook()
sheet = book.active
sheet['A1'] = 1
sheet.cell(row=2, column=2).value = 2
book.save('write2cell.xlsx')


"""
INGRESAR DATOS AL ARCHIVO
"""
book = Workbook()
sheet = book.active

rows = (
    (88, 46, 57),
    (89, 38, 12),
    (23, 59, 78),
    (56, 21, 98),
    (24, 18, 43),
    (34, 15, 67)
)

for row in rows:
    sheet.append(row)

book.save('appending.xlsx')

"""
LEER CELDAS 
"""
book = openpyxl.load_workbook('sample.xlsx')

sheet = book.active

a1 = sheet['A1']
a2 = sheet['A2']
a3 = sheet.cell(row=3, column=1)
print(a1.value)
print(a2.value)
print(a3.value)


""" COMBINAR CELDAS"""
book = Workbook()
sheet = book.active

sheet.merge_cells('A1:B2')

cell = sheet.cell(row=1, column=1)
cell.value = 'Sunny day'
cell.alignment = Alignment(horizontal='center', vertical='center')

book.save('merging.xlsx')

"""FREEZE CELDAD """
book = Workbook()
sheet = book.active

sheet.freeze_panes = 'B2'

book.save('freezing.xlsx')

"""FORMULAS """

book = Workbook()
sheet = book.active

rows = (
    (34, 26),
    (88, 36),
    (24, 29),
    (15, 22),
    (56, 13),
    (76, 18)
)

for row in rows:
    sheet.append(row)

cell = sheet.cell(row=7, column=3)
cell.value = "=SUM(A1:B6)"
cell.font = cell.font.copy(bold=True)

book.save('formulas.xlsx')